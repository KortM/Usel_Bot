import os
import time
from bd import Session, Usel, engine, Base, User_session, User, AccessUsers
import openpyxl
import schedule 
import threading
from datetime import date
from server_socket import main

class Worker:

    def __init__(self) -> None:
        self.filename = 'baza.xlsx'

    def upload_file(self) -> str:
        '''
            Открываем файл с узлами и пишем в БД
        '''
        basedir =os.path.abspath(os.path.dirname(__file__))
        if os.path.exists(f'{basedir}/{self.filename}'):
            session = Session()
            wookbook = openpyxl.load_workbook(self.filename)
            sheet = wookbook.active
            count = 0
            for i in range(0, sheet.max_row):
                tmp = []
                for col in sheet.iter_cols(1, sheet.max_column):
                    tmp.append(str(col[i].value))
                usel = Usel(description = ''.join(tmp))
                session.add(usel)
                session.commit()
                count +=1
                if sheet.max_row % count == 0:
                    print('Обновление: {}%'.format((count / sheet.max_row * 100)))
        else:
            print('Обновление не удалось, файл отсутсвует!')

    def find_data(self, name: str) -> list:
        '''
            Ещем узел в БД по описанию, через LIKE SQL
        '''
        session = Session()
        result = session.query(Usel).filter(Usel.description.like('%{}%'.format(name))).all()
        if result and (result[0].lenght() > 1):
            return result
        else:
            return ['Узел не найден!']
    
    def drop_data(self) -> str:
        '''
            Перед запуском убираем БД и создаем новую, пустую. 
        '''
        try:
            session = Session()
            session.query(Usel).delete()
            session.commit()
            #Base.metadata.create_all(engine)
        except Exception:
            print('Не удалось очистить БД узлов')

    def start(self) -> None:
        '''
            Запуск:
            1. Указываем путь к файлу,
            2. Убираем базу, если она есть и создаем новую,
            3. Заполняем БД. 
        '''
        self.drop_data()
        self.upload_file()
        thread = threading.Thread(target=self.auto_clean_user_sessions, args=())
        thread.start()
        server_thread = threading.Thread(target=main, args=())
        server_thread.start()

    def auto_clean_user_sessions(self):
        '''
            Каждый месяц удаляем активные сессии
        '''
        schedule.every().day.at("00:00").do(self.clear_user_sessions)
        while True:
            schedule.run_pending()
            time.sleep(1)

    def clear_user_sessions(self):
        '''
            Чистим сессии
        '''
        if date.today().day == 1:
            session = User_session()
            try:
                session.query(AccessUsers).delete()
                session.commit()
            except Exception as e:
                session.rollback()

    def is_auth(self, chat_id: str) -> bool:
        '''
            Проверяем на авторизацию.
        '''
        session = User_session()
        user = session.query(AccessUsers).filter_by(chat_id = chat_id).first()
        if user:
            return True
        return False

    def grant_access(self, name: str, password: str, chat_id: str) -> bool:
        '''
            Проверяем есть ли chat_id в таблице уже получивших доступ, если да, то возвращаем True.
            Если нет, то проверяем имя пользователя и пароль, если совпадают, то добавляем chat_id в БД
            и возвращаем True.
            Иначе возвращаем False.
        '''
        session = User_session()
        user = session.query(AccessUsers).filter_by(chat_id = chat_id).first()
        if user:
            return True
        else:
            user = session.query(User).filter_by(name = name).first()
            if user and user.check_password(password):
                new_user = AccessUsers(chat_id=chat_id)
                session.add(new_user)
                session.commit()
                return True
        return False

if __name__ == '__main__':
    w = Worker()
    w.start()